import numpy as np
import h5py
import xlwt
import openpyxl as op

# Read the whole file
# the data too big to wirte in the excel file



# strucutre
f = h5py.File('ff0.hdf5','r')
f.keys()
""" 
by xlwet
# Create the note_book and see it
workbook = xlwt.Workbook()

ws_1 = workbook.add_sheet('Coordinate_1',cell_overwrite_ok=True)
ws_2 = workbook.add_sheet('Coordinate_2',cell_overwrite_ok=True)
ws_3 = workbook.add_sheet('Coordinate_3',cell_overwrite_ok=True)
ws_4 = workbook.add_sheet('Coordinate_4',cell_overwrite_ok=True)
ws_5 = workbook.add_sheet('Coordinate_5',cell_overwrite_ok=True)
ws_6 = workbook.add_sheet('Coordinate_6',cell_overwrite_ok=True)
ws_7 = workbook.add_sheet('Coordinate_7',cell_overwrite_ok=True)
ws_8 = workbook.add_sheet('Coordinate_8',cell_overwrite_ok=True)
"""

"""
by openpyxl
create the note book and see it
"""
workbook = op.Workbook()
ws_1 = workbook.create_sheet('Coordinate_1')
ws_2 = workbook.create_sheet('Coordinate_2')
ws_3 = workbook.create_sheet('Coordinate_3')
ws_4 = workbook.create_sheet('Coordinate_4')
ws_5 = workbook.create_sheet('Coordinate_5')
ws_6 = workbook.create_sheet('Coordinate_6')
ws_7 = workbook.create_sheet('Coordinate_7')
ws_8 = workbook.create_sheet('Coordinate_8')



#for model in f.keys():
    #for yy in f[model].keys():
        #print(yy)

# for geometry part
"""
for model in f.keys():
    for j in list(f['model']['Geometry']):
        print(j)
#
for model in f.keys():
    for j in list(f['model']['Geometry']['geom_part_2']):
        print(j)

for model in f.keys():
    for j in list(f['model']['Geometry']['geom_part_2']['geom_domain_1']):
        print(j)

for model in f.keys():
    for j in list(f['model']['Geometry']['geom_part_2']['geom_domain_1']['coordinates']):
        print(j)
"""
        
for j in range(1,9):
    if j == 1:
        ws = ws_1
        for name in ['geom_part_'+str(j)]:
            print(name,":")
            for k in list(f['model']['Geometry'][name]):
                print(k)
                n =0
                for s in list(f['model']['Geometry'][name][k]['coordinates']):
                    res_coordinates = np.array(s)
                    print(res_coordinates.shape)
                    for index_cor in range(0, np.size(res_coordinates,0)):#np.size(res_coordinates,0)):
                        ws.cell(row =index_cor+1,column = n+1,value = str(res_coordinates[index_cor]))
                    n =n+1
    elif j == 2:
        ws = ws_2
        for name in ['geom_part_'+str(j)]:
            print(name,":")
            for k in list(f['model']['Geometry'][name]):
                print(k)
                n =0
                for s in list(f['model']['Geometry'][name][k]['coordinates']):
                    res_coordinates = np.array(s)
                    print(res_coordinates.shape)
                    for index_cor in range(0,np.size(res_coordinates,0)):
                        #ws.write(index_cor,n,str(res_coordinates[index_cor]))
                        ws.cell(row =index_cor+1, column = n+1 ,value = str(res_coordinates[index_cor]))
                    n =n+1
    elif j ==3:
        ws = ws_3
        for name in ['geom_part_'+str(j)]:
            print(name,":")
            for k in list(f['model']['Geometry'][name]):
                print(k)
                n =0
                for s in list(f['model']['Geometry'][name][k]['coordinates']):
                    res_coordinates = np.array(s)
                    print(res_coordinates.shape)
                    for index_cor in range(0, np.size(res_coordinates,0)):
                        #ws.write(index_cor,n,str(res_coordinates[index_cor]))
                        ws.cell(row =index_cor+1,column = n+1,value = str(res_coordinates[index_cor]))
                    n =n+1
    elif j == 4:
        ws = ws_4
        for name in ['geom_part_'+str(j)]:
            print(name,":")
            for k in list(f['model']['Geometry'][name]):
                print(k)
                n =0
                for s in list(f['model']['Geometry'][name][k]['coordinates']):
                    res_coordinates = np.array(s)
                    print(res_coordinates.shape)
                    for index_cor in range(0, np.size(res_coordinates,0)):
                        #ws.write(index_cor,n,str(res_coordinates[index_cor]))
                        ws.cell(row =index_cor+1,column = n+1,value = str(res_coordinates[index_cor]))
                    n =n+1
    elif j ==5:
        ws = ws_5
        for name in ['geom_part_'+str(j)]:
            print(name,":")
            for k in list(f['model']['Geometry'][name]):
                print(k)
                n =0
                for s in list(f['model']['Geometry'][name][k]['coordinates']):
                    res_coordinates = np.array(s)
                    print(res_coordinates.shape)
                    for index_cor in range(0, np.size(res_coordinates,0)):
                        #ws.write(index_cor,n,str(res_coordinates[index_cor]))
                        ws.cell(row =index_cor+1,column = n+1,value = str(res_coordinates[index_cor]))
                    n =n+1
    elif j == 6:
        ws = ws_6
        for name in ['geom_part_'+str(j)]:
            print(name,":")
            for k in list(f['model']['Geometry'][name]):
                print(k)
                n =0
                for s in list(f['model']['Geometry'][name][k]['coordinates']):
                    res_coordinates = np.array(s)
                    print(res_coordinates.shape)
                    for index_cor in range(0, np.size(res_coordinates,0)):
                        #ws.write(index_cor,n,str(res_coordinates[index_cor]))
                        ws.cell(row =index_cor+1,column = n+1,value = str(res_coordinates[index_cor]))
                    n =n+1
    elif j == 7:
        ws = ws_7
        for name in ['geom_part_'+str(j)]:
            print(name,":")
            for k in list(f['model']['Geometry'][name]):
                print(k)
                n =0
                for s in list(f['model']['Geometry'][name][k]['coordinates']):
                    res_coordinates = np.array(s)
                    print(res_coordinates.shape)
                    for index_cor in range(0, np.size(res_coordinates,0)):
                        #ws.write(index_cor,n,str(res_coordinates[index_cor]))
                        ws.cell(row =index_cor+1,column = n+1,value = str(res_coordinates[index_cor]))
                    n =n+1
    elif j ==8:
        ws = ws_8
        for name in ['geom_part_'+str(j)]:
            print(name,":")
            for k in list(f['model']['Geometry'][name]):
                print(k)
                n =0
                for s in list(f['model']['Geometry'][name][k]['coordinates']):
                    res_coordinates = np.array(s)
                    print(res_coordinates.shape)
                    for index_cor in range(0, np.size(res_coordinates,0)):
                        #ws.write(index_cor,n,str(res_coordinates[index_cor]))
                        ws.cell(row =index_cor+1,column = n+1,value = str(res_coordinates[index_cor]))
                    n =n+1
    else:
        print("geometry is over or failed please check again !! ^_^")

"""
        geo_coordinate = np.array(f['model']['Geometry'][name]['geom_domain_1']['coordinates'])
        #print(geo_coordinate)
        print(geo_coordinate.shape)
        print("---------finished----------")
"""

"""save the coordinate into excel file """
#coordinate = np.array(f['model']['Geometry']['geom_part_1']['geom_domain_1']['coordinates'])
#print(coordinate.shape)

#for i in range(0,3):
    #for j in range(0,np.size(coordinate,1)):
    #num = coordinate[i-1,j]
        #ws.write(j,i,str(coordinate[i,j]))
 
 




print("----------------.----------------------.----------------------------.-------------")
"""
for model in f.keys():
    for j in list(f['model']['Variables']):
        print(j)
"""

#n= 0
for model in f.keys():
    n = 4
    for i in range(1,10):
        print('variable_'+str(i),':')
        #print(list(f['model']['Variables']['variable_'+str(i)]))
        for j in list(f['model']['Variables']['variable_'+str(i)]):
            print(j)

            for num_part in range(1,9):
                if num_part ==1:
                    ws = ws_1
                    #n =4
                    for k in list(f['model']['Variables']['variable_'+str(i)]["var_part_"+str(num_part)]):
                        print(k)
                        #n=4
                        for num_var in list(f['model']['Variables']['variable_'+str(i)]["var_part_"+str(num_part)][k]):
                            print(num_var,":")
                            result = np.array(f['model']['Variables']['variable_'+str(i)]["var_part_"+str(num_part)][k][num_var])
                            print(result.shape)
                            for index_result in range(0,np.size(result,0)):  #np.size(result,0)):
                                #ws.write(index_result,n,str(result[index_result]))
                                ws.cell(row =index_result+1,column = n+1,value = str(result[index_result]))
                            print('write'+str(num_part), 'to exce file :'+str(n),'the variable')
                            #n =n+1
                elif num_part ==2:
                    ws = ws_2
                    #n =4
                    for k in list(f['model']['Variables']['variable_'+str(i)]["var_part_"+str(num_part)]):
                        print(k)
                        #n=4
                        for num_var in list(f['model']['Variables']['variable_'+str(i)]["var_part_"+str(num_part)][k]):
                            print(num_var,":")
                            result = np.array(f['model']['Variables']['variable_'+str(i)]["var_part_"+str(num_part)][k][num_var])
                            print(result.shape)
                            for index_result in range(0,np.size(result,0)):
                                #ws.write(index_result,n,str(result[index_result]))
                                ws.cell(row =index_result+1,column = n+1,value = str(result[index_result]))
                            print('write'+str(num_part), 'to exce file :'+str(n),'the variable')
                            #n =n+1
                elif num_part ==3:
                    ws = ws_3
                    #n =4
                    for k in list(f['model']['Variables']['variable_'+str(i)]["var_part_"+str(num_part)]):
                        print(k)
                        #n=4
                        for num_var in list(f['model']['Variables']['variable_'+str(i)]["var_part_"+str(num_part)][k]):
                            print(num_var,":")
                            result = np.array(f['model']['Variables']['variable_'+str(i)]["var_part_"+str(num_part)][k][num_var])
                            print(result.shape)
                            for index_result in range(0,np.size(result,0)):
                                #ws.write(index_result,n,str(result[index_result]))
                                ws.cell(row =index_result+1,column = n+1,value = str(result[index_result]))
                            print('write'+str(num_part), 'to exce file :'+str(n),'the variable')
                            #n =n+1
                elif num_part ==4:
                    ws = ws_4
                    #n =4
                    for k in list(f['model']['Variables']['variable_'+str(i)]["var_part_"+str(num_part)]):
                        print(k)
                        #n=4
                        for num_var in list(f['model']['Variables']['variable_'+str(i)]["var_part_"+str(num_part)][k]):
                            print(num_var,":")
                            result = np.array(f['model']['Variables']['variable_'+str(i)]["var_part_"+str(num_part)][k][num_var])
                            print(result.shape)
                            for index_result in range(0,np.size(result,0)):
                                #ws.write(index_result,n,str(result[index_result]))
                                ws.cell(row =index_result+1,column = n+1,value = str(result[index_result]))
                            print('write'+str(num_part), 'to exce file :'+str(n),'the variable')
                            #n =n+1
                elif num_part ==5:
                    ws = ws_5
                    #n =4
                    for k in list(f['model']['Variables']['variable_'+str(i)]["var_part_"+str(num_part)]):
                        print(k)
                        #n=4
                        for num_var in list(f['model']['Variables']['variable_'+str(i)]["var_part_"+str(num_part)][k]):
                            print(num_var,":")
                            result = np.array(f['model']['Variables']['variable_'+str(i)]["var_part_"+str(num_part)][k][num_var])
                            print(result.shape)
                            for index_result in range(0,np.size(result,0)):
                                #ws.write(index_result,n,str(result[index_result]))
                                ws.cell(row =index_result+1,column = n+1,value = str(result[index_result]))
                            print('write'+str(num_part), 'to exce file :'+str(n),'the variable')
                            #n =n+1
                elif num_part ==6:
                    ws = ws_6
                    #n =4
                    for k in list(f['model']['Variables']['variable_'+str(i)]["var_part_"+str(num_part)]):
                        print(k)
                        #n=4
                        for num_var in list(f['model']['Variables']['variable_'+str(i)]["var_part_"+str(num_part)][k]):
                            print(num_var,":")
                            result = np.array(f['model']['Variables']['variable_'+str(i)]["var_part_"+str(num_part)][k][num_var])
                            print(result.shape)
                            for index_result in range(0,np.size(result,0)):
                                #ws.write(index_result,n,str(result[index_result]))
                                ws.cell(row =index_result+1,column = n+1,value = str(result[index_result]))
                            print('write'+str(num_part), 'to exce file :'+str(n),'the variable')
                            #n =n+1
                elif num_part ==7:
                    ws = ws_7
                    #n =4
                    for k in list(f['model']['Variables']['variable_'+str(i)]["var_part_"+str(num_part)]):
                        print(k)
                        #n=4
                        for num_var in list(f['model']['Variables']['variable_'+str(i)]["var_part_"+str(num_part)][k]):
                            print(num_var,":")
                            result = np.array(f['model']['Variables']['variable_'+str(i)]["var_part_"+str(num_part)][k][num_var])
                            print(result.shape)
                            for index_result in range(0,np.size(result,0)):
                                #ws.write(index_result,n,str(result[index_result]))
                                ws.cell(row =index_result+1,column = n+1,value = str(result[index_result]))
                            print('write'+str(num_part), 'to exce file :'+str(n),'the variable')
                            #n =n+1
                elif num_part ==8:
                    ws = ws_8
                    #n =4
                    for k in list(f['model']['Variables']['variable_'+str(i)]["var_part_"+str(num_part)]):
                        print(k)
                        #n=4
                        for num_var in list(f['model']['Variables']['variable_'+str(i)]["var_part_"+str(num_part)][k]):
                            print(num_var,":")
                            result = np.array(f['model']['Variables']['variable_'+str(i)]["var_part_"+str(num_part)][k][num_var])
                            print(result.shape)
                            for index_result in range(0,np.size(result,0)):
                                #ws.write(index_result,n,str(result[index_result]))
                                ws.cell(row =index_result+1,column = n+1,value = str(result[index_result]))
                            print('write'+str(num_part), 'to exce file :'+str(n),'the variable')
                            #n =n+1
                else:
                    print("No_result")
        n =n+1
#print(n)

workbook.save('file_six.xlsx')

"""
            for k in list(f['model']['Variables']['variable_'+str(i)]['var_part_1']['var_domain_1']):
                print(k)
                #print(list(f['model']['Variables']['variable_'+str(i)]['var_part_1']['var_domain_1']['tria3']))
                print(f['model']['Variables']['variable_'+str(i)]['var_part_1']['var_domain_1']['tria3'])
                print(list(f['model']['Variables']['variable_'+str(i)]['var_part_1']['var_domain_1']['tria3']))
                #print('p1213')
                #print("..................................................")
"""
